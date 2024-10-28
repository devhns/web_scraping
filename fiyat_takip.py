from bs4 import BeautifulSoup
import requests
from openpyxl import load_workbook

url = 'https://www.trendyol.com/apple/iphone-14-128-gb-gece-yarisi-p-355707175'

wpage = requests.get(url)
html_wpage = BeautifulSoup(wpage.content, "html.parser")
name = html_wpage.find ("h1", class_="pr-new-br").getText()
price = html_wpage.find ("span", class_="prc-dsc").getText()

from openpyxl import load_workbook
workbook = load_workbook("fiyat_takip.xlsx")
sheet = workbook["Sayfa1"]
max_row = sheet.max_row
sheet[f'A{max_row+1}'] = name
sheet[f'B{max_row+1}'] = price

workbook.save("fiyat_takip.xlsx")